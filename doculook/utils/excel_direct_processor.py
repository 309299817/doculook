# Copyright (c) Opendatalab. All rights reserved.
"""
Excel文档直接处理器
直接提取Excel文档中的数据、格式和图片，不经过PDF转换
"""

import io
import json
from pathlib import Path
from typing import Dict, List, Union, Any
from loguru import logger

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Fill, Border, Alignment
except ImportError:
    logger.warning("openpyxl 未安装，Excel文档支持将被禁用")
    openpyxl = None
    load_workbook = None

try:
    import pandas as pd
except ImportError:
    logger.warning("pandas 未安装，Excel文档支持将被禁用")
    pd = None

try:
    from PIL import Image, ImageEnhance
except ImportError:
    logger.warning("PIL (Pillow) 未安装，图片处理功能将受限")
    Image = None
    ImageEnhance = None


def _format_cell_value(cell, preserve_formula=True):
    """
    格式化单元格值，保留原始类型和格式，处理Markdown冲突
    
    Args:
        cell: openpyxl单元格对象
        preserve_formula: 是否保留公式
        
    Returns:
        str: 格式化后的值
    """
    if cell.value is None:
        return ''
    
    # 如果是公式且要保留公式
    if preserve_formula and str(cell.value).startswith('='):
        formula_text = f"{cell.value} (结果: {cell.displayed_value or ''})"
        return _escape_markdown_conflicts(formula_text)
    
    # 处理不同数据类型
    if cell.data_type == 'n':  # 数字
        if cell.number_format:
            try:
                # 尝试使用Excel的数字格式
                value = str(cell.displayed_value or cell.value)
            except:
                value = str(cell.value)
        else:
            value = str(cell.value)
        return _escape_markdown_conflicts(value)
    elif cell.data_type == 'd':  # 日期
        return _escape_markdown_conflicts(str(cell.value))
    elif cell.data_type == 'b':  # 布尔值
        return 'TRUE' if cell.value else 'FALSE'
    else:
        # 处理字符串类型，可能包含换行符和特殊字符
        value = str(cell.value)
        return _escape_markdown_conflicts(value)


def _escape_markdown_conflicts(text, table_format="markdown"):
    """
    转义Markdown格式冲突的字符，支持多种输出格式
    
    Args:
        text: 原始文本
        table_format: 表格格式 ("markdown", "html", "github")
        
    Returns:
        str: 转义后的文本
    """
    if not text:
        return ''
    
    # 根据不同格式处理换行符
    if table_format == "html":
        # HTML表格：使用<br/>标签
        text = text.replace('\n', '<br/>')
        text = text.replace('\r\n', '<br/>')
        text = text.replace('\r', '<br/>')
    elif table_format == "github":
        # GitHub风格：支持<br/>的Markdown
        text = text.replace('\n', '<br/>')
        text = text.replace('\r\n', '<br/>')
        text = text.replace('\r', '<br/>')
    else:
        # 标准Markdown：替换为空格+双空格实现软换行
        text = text.replace('\n', '  ')
        text = text.replace('\r\n', '  ')
        text = text.replace('\r', '  ')
    
    # 转义表格分隔符（所有格式都需要）
    text = text.replace('|', '&#124;')
    
    # 转义其他可能冲突的字符（保守处理）
    text = text.replace('*', '\\*')  # 转义斜体/粗体标记
    text = text.replace('_', '\\_')  # 转义下划线
    text = text.replace('`', '\\`')  # 转义代码标记
    
    # 处理连续的空格（保持格式）
    text = text.replace('  ', ' &nbsp;')
    
    return text


def _create_html_table(table_data, merged_cells=None):
    """
    创建HTML表格（完全支持合并单元格和换行）
    
    Args:
        table_data: 表格数据
        merged_cells: 合并单元格信息
        
    Returns:
        str: HTML表格代码
    """
    if not table_data:
        return ""
    
    html_lines = ['<table border="1">']
    
    for row_idx, row in enumerate(table_data):
        html_lines.append('  <tr>')
        
        for col_idx, cell_value in enumerate(row):
            cell_key = f"{row_idx+1}_{col_idx+1}"
            
            # 检查是否是被合并的单元格（跳过）
            if merged_cells and cell_key in merged_cells and merged_cells[cell_key].get('is_merged_cell'):
                continue
            
            # 构建单元格标签
            cell_tag = 'th' if row_idx == 0 else 'td'
            cell_attrs = []
            
            # 添加合并属性
            if merged_cells and cell_key in merged_cells and merged_cells[cell_key].get('is_merged'):
                merge_info = merged_cells[cell_key]
                if merge_info.get('colspan', 1) > 1:
                    cell_attrs.append(f'colspan="{merge_info["colspan"]}"')
                if merge_info.get('rowspan', 1) > 1:
                    cell_attrs.append(f'rowspan="{merge_info["rowspan"]}"')
            
            attrs_str = ' ' + ' '.join(cell_attrs) if cell_attrs else ''
            
            # 处理单元格内容（支持HTML换行）
            processed_value = _escape_markdown_conflicts(str(cell_value), table_format="html")
            
            # 确保图片引用在HTML中也能被正确识别和替换
            # 保持原始的图片引用格式，以便后续OCR结果替换
            if '![图片' in str(cell_value) and '](images/excel_image_' in str(cell_value):
                # 这是图片引用，保持原格式以便后续替换
                processed_value = str(cell_value)  # 不进行转义，保持原始格式
            
            html_lines.append(f'    <{cell_tag}{attrs_str}>{processed_value}</{cell_tag}>')
        
        html_lines.append('  </tr>')
    
    html_lines.append('</table>')
    return '\n'.join(html_lines)


def _create_enhanced_markdown_table(table_data, merged_cells=None):
    """
    创建增强的Markdown表格（支持GitHub风格的<br/>换行）
    
    Args:
        table_data: 表格数据
        merged_cells: 合并单元格信息
        
    Returns:
        str: 增强Markdown表格
    """
    if not table_data:
        return ""
    
    markdown_lines = []
    
    # 添加说明注释
    markdown_lines.append("<!-- 增强表格：支持换行和合并单元格 -->")
    
    for row_idx, row in enumerate(table_data):
        # 处理每个单元格
        processed_row = []
        for col_idx, cell_value in enumerate(row):
            cell_key = f"{row_idx+1}_{col_idx+1}"
            
            # 被合并的单元格留空
            if merged_cells and cell_key in merged_cells and merged_cells[cell_key].get('is_merged_cell'):
                processed_row.append('')
                continue
            
            # 处理单元格内容
            processed_value = _escape_markdown_conflicts(str(cell_value), table_format="github")
            
            # 确保图片引用在增强Markdown中也能被正确识别和替换
            if '![图片' in str(cell_value) and '](images/excel_image_' in str(cell_value):
                # 这是图片引用，保持原格式以便后续替换
                processed_value = str(cell_value)  # 不进行转义，保持原始格式
            
            # 添加合并信息注释
            if merged_cells and cell_key in merged_cells and merged_cells[cell_key].get('is_merged'):
                merge_info = merged_cells[cell_key]
                if merge_info.get('colspan', 1) > 1 or merge_info.get('rowspan', 1) > 1:
                    processed_value += f" <!-- 合并: {merge_info.get('colspan', 1)}x{merge_info.get('rowspan', 1)} -->"
            
            processed_row.append(processed_value)
        
        # 构建Markdown行
        row_text = "| " + " | ".join(processed_row) + " |"
        markdown_lines.append(row_text)
        
        # 添加表头分隔符
        if row_idx == 0:
            separator = "| " + " | ".join(["---"] * len(processed_row)) + " |"
            markdown_lines.append(separator)
    
    return '\n'.join(markdown_lines)


def _get_cell_style_info(cell):
    """
    获取单元格样式信息
    
    Args:
        cell: openpyxl单元格对象
        
    Returns:
        dict: 样式信息
    """
    style_info = {}
    
    # 字体信息
    if cell.font:
        if cell.font.bold:
            style_info['bold'] = True
        if cell.font.italic:
            style_info['italic'] = True
        if cell.font.underline:
            style_info['underline'] = True
        if cell.font.color and hasattr(cell.font.color, 'rgb'):
            style_info['font_color'] = cell.font.color.rgb
    
    # 填充色
    if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color.rgb != '00000000':
        style_info['bg_color'] = cell.fill.start_color.rgb
    
    # 对齐方式
    if cell.alignment:
        if cell.alignment.horizontal:
            style_info['align'] = cell.alignment.horizontal
        if cell.alignment.vertical:
            style_info['valign'] = cell.alignment.vertical
    
    return style_info


def _extract_images_from_excel_worksheet(worksheet):
    """
    从Excel工作表中提取图片及其位置信息
    
    Args:
        worksheet: openpyxl工作表对象
        
    Returns:
        list: 图片信息列表，包含位置和数据
    """
    images = []
    
    try:
        if hasattr(worksheet, '_images'):
            for img in worksheet._images:
                image_info = {
                    'data': None,
                    'row': None,
                    'col': None,
                    'width': None,
                    'height': None
                }
                
                # 获取图片数据
                if hasattr(img, '_data'):
                    image_info['data'] = img._data()
                
                # 获取图片位置
                if hasattr(img, 'anchor'):
                    anchor = img.anchor
                    if hasattr(anchor, '_from'):
                        image_info['row'] = anchor._from.row + 1  # Excel行从1开始
                        image_info['col'] = anchor._from.col + 1  # Excel列从1开始
                
                # 获取图片尺寸
                if hasattr(img, 'width'):
                    image_info['width'] = img.width
                if hasattr(img, 'height'):
                    image_info['height'] = img.height
                
                if image_info['data']:
                    images.append(image_info)
                    
    except Exception as e:
        logger.debug(f"提取工作表图片失败: {e}")
    
    return images


def _process_merged_cells(worksheet):
    """
    处理合并单元格信息
    
    Args:
        worksheet: openpyxl工作表对象
        
    Returns:
        dict: 合并单元格信息
    """
    merged_cells = {}
    
    try:
        for merged_range in worksheet.merged_cells.ranges:
            min_row, min_col, max_row, max_col = (
                merged_range.min_row, merged_range.min_col,
                merged_range.max_row, merged_range.max_col
            )
            
            # 记录合并区域
            merged_cells[f"{min_row}_{min_col}"] = {
                'rowspan': max_row - min_row + 1,
                'colspan': max_col - min_col + 1,
                'is_merged': True
            }
            
            # 标记被合并的单元格
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    if row != min_row or col != min_col:
                        merged_cells[f"{row}_{col}"] = {'is_merged_cell': True}
                        
    except Exception as e:
        logger.debug(f"处理合并单元格失败: {e}")
    
    return merged_cells


def process_excel_document_directly(excel_bytes: bytes, table_format: str = "auto") -> Dict:
    """
    直接处理Excel文档，提取数据、格式和图片
    
    Args:
        excel_bytes: Excel文档的字节数据
        table_format: 表格输出格式 ("auto", "markdown", "html", "github")
        
    Returns:
        dict: 包含处理结果的字典
    """
    if load_workbook is None:
        raise ImportError("openpyxl 未安装，无法处理Excel文档")
    
    try:
        # 使用BytesIO读取Excel文档
        excel_stream = io.BytesIO(excel_bytes)
        workbook = load_workbook(excel_stream, data_only=False)  # 保留公式
        
        result = {
            'worksheets': [],
            'images': [],
            'markdown_content': '',
            'has_data': False,
            'has_images': False,
            'worksheet_count': len(workbook.worksheets)
        }
        
        logger.info(f"开始处理Excel文档，包含 {len(workbook.worksheets)} 个工作表")
        
        markdown_parts = []
        all_images = []
        
        # 处理每个工作表
        table_index = 0  # 全局表格计数器
        for sheet_idx, worksheet in enumerate(workbook.worksheets):
            sheet_name = worksheet.title
            logger.info(f"处理工作表: {sheet_name}")
            
            # 工作表标题
            markdown_parts.append(f"## 工作表: {sheet_name}\n")
            
            # 获取工作表的数据范围
            if worksheet.max_row == 1 and worksheet.max_column == 1:
                # 空工作表
                markdown_parts.append("*此工作表为空*\n")
                continue
            
            # 处理合并单元格
            merged_cells = _process_merged_cells(worksheet)
            
            # 提取图片
            sheet_images = _extract_images_from_excel_worksheet(worksheet)
            
            # 构建表格数据，包含位置信息
            table_data = []
            image_positions = {}  # 存储图片在表格中的位置
            
            # 记录图片位置
            for img_idx, img_info in enumerate(sheet_images):
                if img_info['row'] and img_info['col']:
                    pos_key = f"{img_info['row']}_{img_info['col']}"
                    image_positions[pos_key] = {
                        'image_index': len(all_images),
                        'data': img_info['data'],
                        'width': img_info.get('width'),
                        'height': img_info.get('height')
                    }
                    all_images.append(img_info['data'])
            
            # 获取实际使用的范围
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # 构建Markdown表格
            for row_idx in range(1, max_row + 1):
                row_data = []
                
                for col_idx in range(1, max_col + 1):
                    cell_key = f"{row_idx}_{col_idx}"
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    
                    # 检查是否是被合并的单元格
                    if cell_key in merged_cells and merged_cells[cell_key].get('is_merged_cell'):
                        # 被合并的单元格添加空值，保持列数一致
                        row_data.append('')
                        continue
                    
                    # 获取单元格值
                    cell_value = _format_cell_value(cell)
                    
                    # 如果是合并单元格的主单元格，添加合并标记
                    if cell_key in merged_cells and merged_cells[cell_key].get('is_merged'):
                        merge_info = merged_cells[cell_key]
                        if merge_info['rowspan'] > 1 or merge_info['colspan'] > 1:
                            cell_value = f"{cell_value} <span colspan='{merge_info['colspan']}' rowspan='{merge_info['rowspan']}'></span>"
                    
                    # 获取样式信息
                    style_info = _get_cell_style_info(cell)
                    
                    # 应用样式到Markdown
                    if style_info.get('bold'):
                        cell_value = f"**{cell_value}**"
                    if style_info.get('italic'):
                        cell_value = f"*{cell_value}*"
                    
                    # 检查该位置是否有图片
                    if cell_key in image_positions:
                        img_info = image_positions[cell_key]
                        img_ref = f"![图片{img_info['image_index']+1}](images/excel_image_{img_info['image_index']+1}.png)"
                        if cell_value.strip():
                            cell_value = f"{cell_value}<br/>{img_ref}"
                        else:
                            cell_value = img_ref
                    
                    row_data.append(cell_value or '')
                
                if any(cell.strip() for cell in row_data):  # 如果行不为空
                    table_data.append(row_data)
            
            # 生成表格（支持多种格式）
            if table_data:
                result['has_data'] = True
                table_index += 1  # 增加表格计数
                
                # 智能选择表格格式
                has_complex_content = any(
                    '\n' in str(cell) or '\r' in str(cell) or '<br/>' in str(cell) 
                    for row in table_data for cell in row
                )
                has_merged_cells = bool(merged_cells and any(
                    info.get('is_merged') for info in merged_cells.values()
                ))
                
                # 根据用户设置和内容复杂度选择格式
                if table_format == "html" or (table_format == "auto" and (has_complex_content or has_merged_cells)):
                    # HTML格式：完全支持换行和合并单元格
                    markdown_parts.append(f"\n**表格 {table_index}（HTML格式，完全支持）:**\n")
                    html_table = _create_html_table(table_data, merged_cells)
                    markdown_parts.append(html_table)
                    
                elif table_format == "github" or (table_format == "auto" and has_complex_content):
                    # GitHub风格：支持<br/>换行
                    markdown_parts.append(f"\n**表格 {table_index}（GitHub增强格式）:**\n")
                    enhanced_markdown = _create_enhanced_markdown_table(table_data, merged_cells)
                    markdown_parts.append(enhanced_markdown)
                    
                else:
                    # 标准Markdown：兼容性最好
                    markdown_parts.append(f"\n**表格 {table_index}:**\n")
                    
                    if len(table_data) > 0:
                        # 处理每行数据
                        processed_table_data = []
                        for row in table_data:
                            processed_row = []
                            for cell in row:
                                # 检查是否包含图片引用
                                if '![图片' in str(cell) and '](images/excel_image_' in str(cell):
                                    # 图片引用保持原格式，不转义
                                    processed_cell = str(cell)
                                else:
                                    # 普通文本进行转义
                                    processed_cell = _escape_markdown_conflicts(str(cell), table_format="markdown")
                                processed_row.append(processed_cell)
                            processed_table_data.append(processed_row)
                        
                        # 生成标准Markdown表格
                        header_row = "| " + " | ".join(processed_table_data[0]) + " |"
                        separator_row = "| " + " | ".join(["---"] * len(processed_table_data[0])) + " |"
                        markdown_parts.append(header_row)
                        markdown_parts.append(separator_row)
                        
                        # 数据行
                        for row in processed_table_data[1:]:
                            # 确保行长度一致
                            while len(row) < len(processed_table_data[0]):
                                row.append('')
                            data_row = "| " + " | ".join(row) + " |"
                            markdown_parts.append(data_row)
                
                # 如果有复杂内容，添加格式说明
                if has_complex_content or has_merged_cells:
                    if table_format == "auto":
                        markdown_parts.append("<!-- 注意：此表格包含换行或合并单元格，在不同Markdown解析器中显示效果可能不同 -->")
                    if has_complex_content:
                        markdown_parts.append("<!-- 单元格内容包含换行符，已转换为适当格式 -->")
                    if has_merged_cells:
                        markdown_parts.append("<!-- 包含合并单元格，HTML格式中有完整支持 -->")
                
                markdown_parts.append("")  # 空行分隔
            
            # 存储工作表信息（不包含图片字节数据以避免JSON序列化问题）
            worksheet_info = {
                'name': sheet_name,
                'data': table_data,
                'images': [{'row': img.get('row'), 'col': img.get('col'), 'width': img.get('width'), 'height': img.get('height')} for img in sheet_images],  # 只保存位置信息
                'merged_cells': merged_cells,
                'row_count': max_row,
                'col_count': max_col
            }
            result['worksheets'].append(worksheet_info)
        
        # 设置结果
        result['images'] = all_images
        result['has_images'] = len(all_images) > 0
        result['markdown_content'] = "\n".join(markdown_parts)
        result['image_count'] = len(all_images)
        
        logger.info(f"Excel处理完成: {result['worksheet_count']} 个工作表, {result['image_count']} 张图片")
        
        return result
            
    except Exception as e:
        logger.error(f"Excel直接处理失败: {e}")
        raise


if __name__ == "__main__":
    # 测试代码
    logger.info("Excel直接处理器加载完成")
